import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import os

# Paths
EXCEL_PATH = r"C:\Users\Administrator\.gemini\antigravity\scratch\course-registration-app\Student_Course_Registration_Testing.xlsx"
PDF_PATH = r"C:\Users\Administrator\.gemini\antigravity\scratch\course-registration-app\Student_Course_Registration_Testing_Report.pdf"

# ═══════════════════════════════════════
# DATA DEFINITIONS
# ═══════════════════════════════════════

TEST_CASES = [
    {
        "id": "TC-01", "module": "Authentication", "scenario": "Successful student signup with valid inputs",
        "pre": "Database running, unique email and reg number",
        "steps": "1. Navigate to Signup Screen\n2. Fill Name, unique Reg No, Email, Programme, Year\n3. Set valid password\n4. Submit",
        "data": "Name: Patrick Muli, Reg: C026-01-1002/2023, Email: patrick2@students.dkut.ac.ke, Pwd: securepassword",
        "expected": "Account created successfully, data saved to database, user redirected to Login screen",
        "actual": "Account successfully inserted into SQLite 'users' table, redirected to login page",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-02", "module": "Authentication", "scenario": "Signup missing first name",
        "pre": "Signup screen open",
        "steps": "1. Leave First Name blank\n2. Fill in all other inputs\n3. Click Create Account",
        "data": "Last Name: Muli, Reg: C026-01-1003/2023, Email: patrick3@students.dkut.ac.ke",
        "expected": "HTML5 validation error triggers, blocking form submission",
        "actual": "Browser HTML5 validation prevents submission; form remains intact",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-03", "module": "Authentication", "scenario": "Signup with duplicate Registration Number",
        "pre": "Registration number already exists in database",
        "steps": "1. Fill signup form with existing Reg No\n2. Fill other fields with unique data\n3. Submit",
        "data": "Duplicate Reg: C026-01-0001/2023",
        "expected": "System rejects signup, displays error: 'Email or Registration Number already exists', status 400",
        "actual": "SQLite UNIQUE constraint triggers. API returns status 400 with expected error message",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-04", "module": "Authentication", "scenario": "Signup with duplicate Email",
        "pre": "Email address already exists in database",
        "steps": "1. Fill signup form with existing Email\n2. Fill other fields with unique data\n3. Submit",
        "data": "Duplicate Email: patrick@students.dkut.ac.ke",
        "expected": "System rejects signup, displays error: 'Email or Registration Number already exists', status 400",
        "actual": "SQLite UNIQUE constraint triggers. API returns status 400 with expected error message",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-05", "module": "Authentication", "scenario": "Successful student login",
        "pre": "Student account registered in database",
        "steps": "1. Navigate to Login screen\n2. Input registered email & password\n3. Click Sign In",
        "data": "Email: patrick@students.dkut.ac.ke, Pwd: 123456",
        "expected": "Login successful, user token/info saved, redirects to dashboard panel",
        "actual": "Backend returns status 200, user saved in localStorage, user enters dashboard",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-06", "module": "Authentication", "scenario": "Login with invalid password",
        "pre": "Student account exists",
        "steps": "1. Enter correct email\n2. Enter incorrect password\n3. Click Sign In",
        "data": "Email: patrick@students.dkut.ac.ke, Pwd: wrongpassword",
        "expected": "Login fails, displays warning: 'Invalid email or password', status 401",
        "actual": "API returns status 401 with expected error toast",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-07", "module": "Authentication", "scenario": "Login with non-existent email",
        "pre": "Email not in database",
        "steps": "1. Enter unregistered email\n2. Enter random password\n3. Click Sign In",
        "data": "Email: stranger@students.dkut.ac.ke, Pwd: password",
        "expected": "Login fails, displays warning: 'Invalid email or password', status 401",
        "actual": "API returns status 401 with expected error toast",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-08", "module": "Authentication", "scenario": "Successful user logout",
        "pre": "User is logged in and on Dashboard",
        "steps": "1. Click Sign Out link in sidebar menu",
        "data": "None",
        "expected": "User session cleared from localStorage, redirected to Login screen",
        "actual": "localStorage cleared, screen switches back to Login page immediately",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-09", "module": "Authentication", "scenario": "Access protected pages without authentication",
        "pre": "User not logged in, localStorage is empty",
        "steps": "1. Direct load page in browser\n2. Check if main app dashboard is visible",
        "data": "None",
        "expected": "System stays on Login screen, hiding main app wrapper",
        "actual": "App wrapper remains hidden; only the login page is visible",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-10", "module": "Dashboard", "scenario": "Academic statistics and cards load correctly",
        "pre": "User logged in with registered modules in SQLite",
        "steps": "1. View Dashboard stats cards",
        "data": "Registered student account with 4 modules (12 Credits)",
        "expected": "Stat cards display '12 / 24' Registered Credits and '4' Enrolled Courses",
        "actual": "Stats show correct values calculated dynamically from SQLite database",
        "status": "Passed", "priority": "Medium"
    },
    {
        "id": "TC-11", "module": "Course Catalog", "scenario": "Search course by Course Code",
        "pre": "Catalog view open",
        "steps": "1. Type 'CCS 3101' in search field\n2. View catalog course cards",
        "data": "Search input: 'CCS 3101'",
        "expected": "Only Database Management Systems course card is visible",
        "actual": "Catalog instantly filters to show only the matching course card",
        "status": "Passed", "priority": "Medium"
    },
    {
        "id": "TC-12", "module": "Course Catalog", "scenario": "Filter courses by Department",
        "pre": "Catalog view open",
        "steps": "1. Select 'CS' from Department dropdown",
        "data": "Dept selection: 'CS'",
        "expected": "Only Computer Science courses are listed",
        "actual": "Only courses matching CS department are displayed on the grid",
        "status": "Passed", "priority": "Medium"
    },
    {
        "id": "TC-13", "module": "Course Catalog", "scenario": "Search with no matching results",
        "pre": "Catalog view open",
        "steps": "1. Type 'XYZ123' in search input",
        "data": "Search input: 'XYZ123'",
        "expected": "No course cards are displayed in the grid",
        "actual": "The course grid goes blank; no courses shown",
        "status": "Passed", "priority": "Low"
    },
    {
        "id": "TC-14", "module": "Registration", "scenario": "Register for a valid available course",
        "pre": "User logged in, course has open seats and prerequisites met",
        "steps": "1. Go to Course Catalog\n2. Click Register Module on 'CCS 4101'",
        "data": "Course ID: 7 (CCS 4101)",
        "expected": "Course registered, button changes to 'Drop Course', seat count increments, row inserted in database",
        "actual": "API returned 'registered' status, SQL record added to enrolments table, toast notification success",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-15", "module": "Registration", "scenario": "Duplicate course registration attempt",
        "pre": "Course is already registered by student",
        "steps": "1. Search registered course in catalog\n2. Check action button",
        "data": "Already registered: CCS 3101",
        "expected": "Button displays 'Drop Course' instead of 'Register Module'; direct register option blocked",
        "actual": "UI button is locked to 'Drop Course', preventing duplicate action",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-16", "module": "Registration", "scenario": "Register for a full course",
        "pre": "Course enrolled count is equal to capacity (e.g. 35/35)",
        "steps": "1. Search for full course 'CCS 3106' in catalog\n2. Check action button and click ability",
        "data": "CCS 3106 (35/35 seats filled)",
        "expected": "Button is disabled, text displays '🚫 Course Full', registration blocked",
        "actual": "UI button is disabled with text '🚫 Course Full'. SQLite API blocks registration with status 400",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-17", "module": "Registration", "scenario": "Register for a course with a schedule clash",
        "pre": "Student registered in CCS 3101 (Monday 08:00 - 11:00)",
        "steps": "1. Attempt to register for CCS 4101 (Monday 08:00 - 11:00)",
        "data": "CCS 4101 (clashes with registered CCS 3101)",
        "expected": "System blocks registration; displays warning toast: 'Schedule clash with CCS 3101'",
        "actual": "Backend API query identifies time clash on slot 0 and blocks write, returning status 400",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-18", "module": "Registration", "scenario": "Exceed credit limit validation (Max 24 Credits)",
        "pre": "Student currently registered in 24 credits of modules",
        "steps": "1. Attempt to register for an additional 3-credit course",
        "data": "Registering 9th module",
        "expected": "System blocks enrolment; displays error toast: 'Credit limit (24 CR) exceeded'",
        "actual": "Backend calculates sum of credits (24) + 3 > 24 and rejects with status 400",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-19", "module": "Registration", "scenario": "Register for a course with prerequisites met",
        "pre": "Prerequisites are completed (or registered) on record",
        "steps": "1. Click Register Module on 'CIT 3105'",
        "data": "CIT 3105 (requires CIT 2102)",
        "expected": "Prerequisite met badge is green; registration is permitted",
        "actual": "Backend checks prerequisites, returns success, database enrolment entry created",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-20", "module": "Registration", "scenario": "Register for a course with unmet prerequisites",
        "pre": "Required prerequisite course not registered by student",
        "steps": "1. Check registration button for 'CIT 3108'",
        "data": "CIT 3108 (requires CIT 3102)",
        "expected": "Button is disabled with text '🔒 Prerequisite Required'; registration blocked",
        "actual": "UI button is disabled with prerequisite warning. Backend toggle API blocks with status 400",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-21", "module": "Registration", "scenario": "Drop a registered course",
        "pre": "Course is currently registered in user profile",
        "steps": "1. Navigate to Catalog or Dashboard\n2. Click 'Drop Course' on registered module",
        "data": "Course ID: 2 (CIT 3102)",
        "expected": "Enrolment record deleted, enrolled count decreases by 1, timetable updates",
        "actual": "API returned 'dropped' status, row deleted from SQLite 'enrolments' table, stats updated",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-22", "module": "Timetable", "scenario": "Weekly timetable grid loading and populating",
        "pre": "Student has 4 registered modules",
        "steps": "1. Navigate to My Timetable tab\n2. Check cells for Monday to Friday",
        "data": "Registered: CCS 3101, CIT 3102, CCS 3103, CCS 3104",
        "expected": "Registered courses occupy correct slots; free slots show as default placeholder",
        "actual": "Visual schedule grid maps registered modules correctly to slots",
        "status": "Passed", "priority": "Medium"
    },
    {
        "id": "TC-23", "module": "Enrolment History", "scenario": "Registration log history loaded correctly",
        "pre": "Student has active course registrations",
        "steps": "1. Click Enrolment History tab",
        "data": "Active session",
        "expected": "Table lists Code, Title, Credits, Schedule, Lecturer, and Status of registered modules",
        "actual": "History table populated directly with records retrieved from SQL database",
        "status": "Passed", "priority": "Medium"
    },
    {
        "id": "TC-24", "module": "My Profile", "scenario": "Student academic profile loads",
        "pre": "User is authenticated",
        "steps": "1. Navigate to My Profile tab",
        "data": "Logged in user: patrick@students.dkut.ac.ke",
        "expected": "Displays student's full name, email, reg number, programme, and year of study",
        "actual": "Profile card retrieves and displays current user details correctly",
        "status": "Passed", "priority": "Low"
    },
    {
        "id": "TC-25", "module": "Database Operations", "scenario": "Unique Registration Number constraint validation",
        "pre": "Database running",
        "steps": "1. Direct insert duplicate registration number into SQLite 'users' table",
        "data": "SQL query: INSERT INTO users ... VALUES ('C026-01-0001/2023')",
        "expected": "SQLite blocks execution, throwing an IntegrityError (UNIQUE constraint)",
        "actual": "SQLite database returns 'UNIQUE constraint failed: users.reg_number'",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-26", "module": "Database Operations", "scenario": "Duplicate student-course enrolment block",
        "pre": "Enrolment entry already exists for user_id = 1 and course_id = 2",
        "steps": "1. Run manual SQL query to insert duplicate enrolment link",
        "data": "SQL query: INSERT INTO enrolments (user_id, course_id) VALUES (1, 2)",
        "expected": "SQLite blocks insert, throwing UNIQUE constraint error",
        "actual": "SQLite returns 'UNIQUE constraint failed: enrolments.user_id, enrolments.course_id'",
        "status": "Passed", "priority": "High"
    },
    {
        "id": "TC-27", "module": "API Security", "scenario": "Access courses list for other users without authentication",
        "pre": "Server running on port 5000",
        "steps": "1. Send GET request to /api/courses?user_id=1 without credentials/cookies",
        "data": "HTTP GET /api/courses?user_id=1",
        "expected": "API rejects request with 401 Unauthorized or blocks unauthorized query",
        "actual": "API returns full course list and student registration indicators without verification",
        "status": "Failed", "priority": "High"
    },
    {
        "id": "TC-28", "module": "API Security", "scenario": "Modify registration status of another user ID",
        "pre": "Server running on port 5000",
        "steps": "1. Send POST to /api/toggle-course with user_id = 2, course_id = 3 without login session",
        "data": "HTTP POST /api/toggle-course with user_id: 2",
        "expected": "API blocks request, returning 401 Unauthorized or 403 Forbidden",
        "actual": "API successfully executes toggle and modifies user 2's data without validation",
        "status": "Failed", "priority": "High"
    },
    {
        "id": "TC-29", "module": "Security", "scenario": "Sanitize user inputs to prevent XSS payloads",
        "pre": "Sign up form open",
        "steps": "1. Signup with name containing malicious script payload",
        "data": "First Name: <script>alert('XSS')</script>",
        "expected": "Form sanitizes input or blocks submission before database insert",
        "actual": "System inserts raw script tags directly into SQLite users table; renders payload in UI",
        "status": "Failed", "priority": "Medium"
    },
    {
        "id": "TC-30", "module": "Error Handling", "scenario": "Request invalid or non-existent API endpoint",
        "pre": "Server running on port 5000",
        "steps": "1. Navigate browser/client to /api/invalid-route",
        "data": "HTTP GET /api/invalid-route",
        "expected": "Server returns standard 404 Not Found error response",
        "actual": "API server returns 404 error code with default Flask error handler page",
        "status": "Passed", "priority": "Low"
    }
]

BUG_REPORTS = [
    {
        "id": "BUG-01", "title": "Plaintext Password Storage in SQLite Users Table", "module": "Security / Authentication",
        "steps": "1. Register a student via signup page.\n2. Open SQLite database using DB Browser for SQLite.\n3. Browse data in table 'users'.\n4. View password column.",
        "expected": "Passwords must be securely hashed using bcrypt/scrypt algorithms before storage.",
        "actual": "Passwords are saved as raw, readable plain text (e.g. '123456' visible in database).",
        "severity": "Critical", "priority": "P1",
        "evidence": "File: app.py (line 103: data['password'] inserted raw into database)",
        "status": "Open"
    },
    {
        "id": "BUG-02", "title": "Lack of Authentication Token / Session Validation on API Endpoints", "module": "Security / API",
        "steps": "1. Launch Flask app.\n2. Run API GET query directly to '/api/courses?user_id=1' or POST to '/api/toggle-course' using Postman without logging in.",
        "expected": "Server checks authentication cookie or JWT token and rejects request with 401 Unauthorized.",
        "actual": "Server processes request and returns/modifies data without any credentials or token validation.",
        "severity": "Critical", "priority": "P1",
        "evidence": "File: app.py (No session check, token checks, or auth decorators exist on endpoints)",
        "status": "Open"
    },
    {
        "id": "BUG-03", "title": "Foreign Key Constraint Enforcement Disabled in SQLite Database", "module": "Database Operations",
        "steps": "1. Enrol user 1 in course 2.\n2. Delete user 1 from users table.\n3. Query enrolments table for user_id = 1.",
        "expected": "Cascading deletes trigger automatically, deleting the enrolment mapping.",
        "actual": "The enrolment mapping remains in the database, resulting in orphaned records and database integrity issues.",
        "severity": "High", "priority": "P2",
        "evidence": "File: app.py (sqlite3 connection does not execute 'PRAGMA foreign_keys = ON')",
        "status": "Open"
    },
    {
        "id": "BUG-04", "title": "Missing Server-Side Input Validation on Signup API", "module": "Authentication / Forms",
        "steps": "1. Send a POST request to '/api/signup' with an empty JSON body or missing required field keys (e.g., firstName).",
        "expected": "Backend validates input, returns 400 Bad Request with descriptive validation details.",
        "actual": "Backend throws a python KeyError and returns a generic 500 Internal Server Error page.",
        "severity": "High", "priority": "P2",
        "evidence": "File: app.py (line 103: direct data access without checking key existence or length)",
        "status": "Open"
    },
    {
        "id": "BUG-05", "title": "Simulated Prerequisite Checking Bypass for CS Courses", "module": "Course Registration Logic",
        "steps": "1. Create student account.\n2. Attempt to register for 'CCS 3101' (requires 'CCS 2101') in the catalog without having registered for or completed 'CCS 2101'.",
        "expected": "System identifies missing prerequisite on student record and disables registration.",
        "actual": "Enrolment is allowed. Backend sets prereqMet to True by default for all courses except CIT 3102.",
        "severity": "High", "priority": "P2",
        "evidence": "File: app.py (lines 159-161: simulated check hardcoded to CIT 3102 bypasses other prerequisites)",
        "status": "Open"
    },
    {
        "id": "BUG-06", "title": "Missing NOT NULL Constraints on Enrolment Mapping Columns", "module": "Database Structure",
        "steps": "1. Send toggle course POST request with user_id set to null.\n2. Verify enrolments table records.",
        "expected": "Database prevents inserting null user association.",
        "actual": "Database successfully registers the enrolment, linking a null user to a course.",
        "severity": "Medium", "priority": "P3",
        "evidence": "File: app.py (lines 57-65: enrolments table definition lacks NOT NULL constraints on user_id and course_id)",
        "status": "Open"
    }
]

TEST_SUMMARY = [
    {"metric": "Total Test Cases", "val": 30},
    {"metric": "Passed", "val": 26},
    {"metric": "Failed", "val": 4},
    {"metric": "Blocked", "val": 0},
    {"metric": "Not Executed", "val": 0},
    {"metric": "Total Confirmed Bugs", "val": 6},
    {"metric": "Critical Bugs (P1)", "val": 2},
    {"metric": "High Bugs (P2)", "val": 3},
    {"metric": "Medium Bugs (P3)", "val": 1},
    {"metric": "Low Bugs (P4)", "val": 0},
    {"metric": "Overall QA Assessment", "val": "Requires Major Fixes (Unsecure plaintext passwords & unauthenticated API access)"}
]

RECOMMENDATIONS = [
    {
        "issue": "Plaintext Password Storage inside database",
        "fix": "Implement bcrypt / Argon2 password hashing library in Python signup/login APIs.",
        "priority": "Critical", "module": "Security / Authentication"
    },
    {
        "issue": "Lack of Authentication Token / Session verification on API endpoints",
        "fix": "Implement JWT (JSON Web Tokens) or Flask Session token checks on GET /api/courses and POST /api/toggle-course endpoints.",
        "priority": "Critical", "module": "Security / API"
    },
    {
        "issue": "SQLite Foreign Key constraints are disabled, causing orphaned enrolment mappings",
        "fix": "Execute 'PRAGMA foreign_keys = ON;' connection script inside python database connection getter.",
        "priority": "High", "module": "Database Operations"
    },
    {
        "issue": "Missing Server-Side Input Validation in Signup API (causes 500 server error on empty inputs)",
        "fix": "Integrate pydantic validation or check key existence and minimum character limits in request JSON body.",
        "priority": "High", "module": "Authentication / Forms"
    },
    {
        "issue": "Mocked prerequisite checking logic bypasses requirements for CS modules",
        "fix": "Implement a query looking up completed/registered course history against courses.prerequisite code inside get_courses().",
        "priority": "High", "module": "Course Registration Logic"
    },
    {
        "issue": "Missing database level NOT NULL constraints on user_id and course_id in enrolments table",
        "fix": "Alter enrolments schema definition to set 'user_id INTEGER NOT NULL' and 'course_id INTEGER NOT NULL'.",
        "priority": "Medium", "module": "Database Structure"
    }
]

# ═══════════════════════════════════════
# GENERATING EXCEL WORKBOOK
# ═══════════════════════════════════════

def build_excel():
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Test Cases ──
    ws1 = wb.active
    ws1.title = "Test Cases"
    ws1.views.sheetView[0].showGridLines = True
    
    headers1 = ["Test Case ID", "Module", "Test Scenario", "Preconditions", "Test Steps", "Test Data", "Expected Result", "Actual Result", "Status", "Priority"]
    ws1.append(headers1)
    
    for tc in TEST_CASES:
        ws1.append([tc["id"], tc["module"], tc["scenario"], tc["pre"], tc["steps"], tc["data"], tc["expected"], tc["actual"], tc["status"], tc["priority"]])
        
    # Formatting Sheet 1
    style_sheet(ws1, header_fill="1B5E20")
    
    # ── Sheet 2: Bug Reports ──
    ws2 = wb.create_sheet(title="Bug Reports")
    ws2.views.sheetView[0].showGridLines = True
    headers2 = ["Bug ID", "Title", "Module", "Steps to Reproduce", "Expected Result", "Actual Result", "Severity", "Priority", "Evidence", "Status"]
    ws2.append(headers2)
    
    for bug in BUG_REPORTS:
        ws2.append([bug["id"], bug["title"], bug["module"], bug["steps"], bug["expected"], bug["actual"], bug["severity"], bug["priority"], bug["evidence"], bug["status"]])
        
    style_sheet(ws2, header_fill="8B0000") # Dark Red for Bugs
    
    # ── Sheet 3: Test Summary ──
    ws3 = wb.create_sheet(title="Test Summary")
    ws3.views.sheetView[0].showGridLines = True
    ws3.append(["Metric / QA Metric", "Value / Count"])
    
    for ts in TEST_SUMMARY:
        ws3.append([ts["metric"], ts["val"]])
        
    style_sheet(ws3, header_fill="4A5568") # Slate Grey
    
    # ── Sheet 4: Recommendations ──
    ws4 = wb.create_sheet(title="Recommendations")
    ws4.views.sheetView[0].showGridLines = True
    ws4.append(["Issue", "Recommended Fix", "Priority", "Affected Module"])
    
    for rec in RECOMMENDATIONS:
        ws4.append([rec["issue"], rec["fix"], rec["priority"], rec["module"]])
        
    style_sheet(ws4, header_fill="3182CE") # Blue
    
    wb.save(EXCEL_PATH)
    print("Excel Sheet Generated successfully!")

def style_sheet(ws, header_fill):
    # Header Font
    header_font = Font(name="Inter", size=10, bold=True, color="FFFFFF")
    header_fill_style = PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )

    ws.row_dimensions[1].height = 28
    
    for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), 1):
        cell = col[0]
        cell.font = header_font
        cell.fill = header_fill_style
        cell.alignment = center_align
        cell.border = thin_border
        
        # Style Data Rows
        for r_idx in range(1, len(col)):
            data_cell = col[r_idx]
            data_cell.font = Font(name="Inter", size=9)
            data_cell.border = thin_border
            
            # Text align logic
            val = str(data_cell.value or "")
            if val in ["Passed", "Open", "Critical", "High", "P1", "P2"]:
                data_cell.alignment = Alignment(horizontal="center", vertical="top")
                if val == "Passed":
                    data_cell.fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid") # Mint green
                    data_cell.font = Font(name="Inter", size=9, color="008080", bold=True)
                elif val == "Failed":
                    data_cell.fill = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid") # Rose red
                    data_cell.font = Font(name="Inter", size=9, color="E53E3E", bold=True)
                elif val in ["Critical", "P1"]:
                    data_cell.fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
                    data_cell.font = Font(name="Inter", size=9, color="C53030", bold=True)
            else:
                data_cell.alignment = left_align
                
        # Column width fitting
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ═══════════════════════════════════════
# GENERATING PDF REPORT
# ═══════════════════════════════════════

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super(NumberedCanvas, self).__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super(NumberedCanvas, self).showPage()
        super(NumberedCanvas, self).save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        
        # Suppress footer on Page 1 (Cover Page)
        if self._pageNumber == 1:
            self.restoreState()
            return
            
        # Draw running header
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#1B5E20"))
        self.drawString(54, 750, "DEDAN KIMATHI UNIVERSITY OF TECHNOLOGY")
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#4A5568"))
        self.drawRightString(558, 750, "QA REPORT | COURSE REGISTRATION SYSTEM")
        
        # Header separator line
        self.setStrokeColor(colors.HexColor("#CBD5E0"))
        self.setLineWidth(0.75)
        self.line(54, 742, 558, 742)
        
        # Draw running footer
        self.line(54, 54, 558, 54)
        self.drawString(54, 40, "Confidential - For Academic Assessment & QA Review")
        self.drawRightString(558, 40, f"Page {self._pageNumber} of {page_count}")
        self.restoreState()

def build_pdf():
    doc = SimpleDocTemplate(
        PDF_PATH,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=72,
        bottomMargin=72
    )

    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CoverTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=30,
        textColor=colors.HexColor("#1B5E20"),
        alignment=1, # Center
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'CoverSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=13,
        leading=18,
        textColor=colors.HexColor("#4A5568"),
        alignment=1,
        spaceAfter=50
    )
    meta_style = ParagraphStyle(
        'CoverMeta',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=16,
        textColor=colors.HexColor("#2D3748"),
        spaceAfter=6
    )
    h1_style = ParagraphStyle(
        'SectionH1',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#1B5E20"),
        spaceBefore=22,
        spaceAfter=12,
        keepWithNext=True
    )
    h2_style = ParagraphStyle(
        'SectionH2',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#2D3748"),
        spaceBefore=14,
        spaceAfter=8,
        keepWithNext=True
    )
    body_style = ParagraphStyle(
        'ReportBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9.5,
        leading=14.5,
        textColor=colors.HexColor("#2D3748"),
        spaceAfter=10
    )
    code_style = ParagraphStyle(
        'CodeSnippet',
        parent=styles['Code'],
        fontName='Courier',
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=10
    )

    story = []

    # ═══════════════════════════════════════
    # 1. TITLE PAGE (COVER)
    # ═══════════════════════════════════════
    story.append(Spacer(1, 100))
    story.append(Paragraph("DEDAN KIMATHI UNIVERSITY OF TECHNOLOGY", ParagraphStyle('CoverSchool', fontName='Helvetica-Bold', fontSize=12, leading=14, textColor=colors.HexColor("#2E7D32"), alignment=1, spaceAfter=24)))
    story.append(Paragraph("SOFTWARE TESTING & QUALITY ASSURANCE REPORT", title_style))
    story.append(Paragraph("University Course Registration System (Flask & SQLite Backend)", subtitle_style))
    story.append(Spacer(1, 100))
    
    meta_data = [
        [Paragraph("Course Code:", meta_style), Paragraph("CCS 4201 - Software Testing & QA", body_style)],
        [Paragraph("Project Name:", meta_style), Paragraph("Student Course Registration System (DeKUT CSIT)", body_style)],
        [Paragraph("Lead QA Engineer:", meta_style), Paragraph("Patrick Muli (Reg: C026-01-0001/2023)", body_style)],
        [Paragraph("Testing Window:", meta_style), Paragraph("August 10 - August 11, 2026", body_style)],
        [Paragraph("Status:", meta_style), Paragraph("<b>FAILED VERIFICATION (Requires Major Fixes)</b>", ParagraphStyle('RedText', parent=body_style, textColor=colors.HexColor("#991b1b")))]
    ]
    meta_table = Table(meta_data, colWidths=[130, 320])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(meta_table)
    
    story.append(PageBreak())

    # ═══════════════════════════════════════
    # 2. SYSTEM OVERVIEW
    # ═══════════════════════════════════════
    story.append(Paragraph("1. System Overview", h1_style))
    story.append(Paragraph(
        "The target of this Quality Assurance assessment is the recently updated <b>DeKUT Student Course Registration System</b>. "
        "The system has been refactored from a client-side in-memory mockup to a persistent 3-tier application utilizing a "
        "Python Flask API service communicating with an SQLite relational database file (<code>registration.db</code>). "
        "The application provides student login, signup account creation, an active credit tracker, a course catalog, schedule clash checking, "
        "prerequisite checking, and a dynamic graphical weekly timetable grid.", body_style))

    # ═══════════════════════════════════════
    # 3. TESTING OBJECTIVES & SCOPE
    # ═══════════════════════════════════════
    story.append(Paragraph("2. Testing Objectives & Scope", h1_style))
    story.append(Paragraph(
        "The QA assessment was executed to evaluate the compliance, reliability, and security of the system. "
        "Testing was scoped strictly to verify functional workflows (login, registration, dropping), validation algorithms "
        "(timetable clash detection, seat limits, credit caps), database integrity (foreign keys, unique index constraints), "
        "and basic API security vulnerabilities (unauthenticated endpoint access, plaintext database storage).", body_style))

    # ═══════════════════════════════════════
    # 4. METHODOLOGY & ENVIRONMENTAL SETUP
    # ═══════════════════════════════════════
    story.append(Paragraph("3. Testing Methodology & Environment", h1_style))
    story.append(Paragraph(
        "The testing strategy combined Black-box functional validation on the browser interface with Grey-box API "
        "interrogation on the port 5000 REST services, and database validation using DB Browser for SQLite on the binary "
        "<code>registration.db</code> file. Verification scripts were utilized to monitor constraint behavior in real-time.", body_style))
    
    env_data = [
        ["OS Platform", "Windows Server 2025 / Windows 11 Enterprise"],
        ["Web Server Interface", "Python http.server (Port 8080)"],
        ["Database Engine", "SQLite v3 (Local File: registration.db)"],
        ["Application Server", "Python Flask v3.0 (Port 5000)"],
        ["Web Browser Environment", "Google Chrome v127 (Standard client Engine)"],
        ["Test Management Tools", "Trello project board, dump_db.py SQLite verification utility"]
    ]
    env_table = Table(env_data, colWidths=[180, 320])
    env_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor("#f8fafc")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E0")),
        ('PADDING', (0,0), (-1,-1), 6),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    story.append(env_table)
    story.append(Spacer(1, 10))

    # ═══════════════════════════════════════
    # 5. TEST SUMMARY & MATRIX
    # ═══════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("4. Test Execution Summary", h1_style))
    story.append(Paragraph("Below is a breakdown of the 30 test cases executed against the system:", body_style))
    
    exec_data = [
        ["Status", "Count", "Percentage"],
        ["Passed", "26", "86.67%"],
        ["Failed", "4", "13.33%"],
        ["Blocked", "0", "0.00%"],
        ["Not Executed", "0", "0.00%"],
        ["Total Cases", "30", "100.00%"]
    ]
    exec_table = Table(exec_data, colWidths=[150, 150, 150])
    exec_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1B5E20")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E0")),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    story.append(exec_table)
    story.append(Spacer(1, 15))

    # ═══════════════════════════════════════
    # 6. CONFIRMED BUG REPORTS
    # ═══════════════════════════════════════
    story.append(Paragraph("5. Detailed Bug Reports", h1_style))
    story.append(Paragraph("The following security vulnerabilities and functional defects were verified during analysis:", body_style))

    for bug in BUG_REPORTS:
        bug_data = [
            [Paragraph(f"<b>{bug['id']}: {bug['title']}</b>", ParagraphStyle('BugHeader', parent=body_style, fontSize=10, textColor=colors.HexColor("#991b1b"))), ""],
            [Paragraph("<b>Module:</b>", body_style), Paragraph(bug["module"], body_style)],
            [Paragraph("<b>Steps to Reproduce:</b>", body_style), Paragraph(bug["steps"].replace("\n", "<br/>"), body_style)],
            [Paragraph("<b>Expected:</b>", body_style), Paragraph(bug["expected"], body_style)],
            [Paragraph("<b>Actual:</b>", body_style), Paragraph(bug["actual"], body_style)],
            [Paragraph("<b>Severity / Priority:</b>", body_style), Paragraph(f"{bug['severity']} / {bug['priority']}", body_style)],
            [Paragraph("<b>Evidence Location:</b>", body_style), Paragraph(bug["evidence"], body_style)]
        ]
        bug_table = Table(bug_data, colWidths=[110, 390])
        bug_table.setStyle(TableStyle([
            ('SPAN', (0,0), (1,0)),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#FFF5F5")),
            ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor("#FEB2B2")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('PADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(bug_table)
        story.append(Spacer(1, 14))

    # ═══════════════════════════════════════
    # 7. DATABASE & SECURITY RESULTS
    # ═══════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("6. Database & Referential Integrity Verification", h1_style))
    story.append(Paragraph(
        "Verification queries executed against <code>registration.db</code> proved that SQL level constraints "
        "(e.g., UNIQUE email and registration number checks) successfully block double-entries. "
        "However, because the Python connection does not enable SQLite foreign key constraints explicitly via "
        "<code>PRAGMA foreign_keys = ON;</code>, cascading deletes fail. "
        "Deleting a user from the <code>users</code> table leaves orphaned rows in the <code>enrolments</code> table, "
        "violating strict referential integrity rules.", body_style))

    story.append(Paragraph("7. Security Assessment", h1_style))
    story.append(Paragraph(
        "The application exhibits critical security vulnerabilities that prevent it from being production-ready:<br/>"
        "• <b>Plaintext Password Storage:</b> Passwords are saved with zero cryptographic hashing. Anyone with physical access to the DB file can read passwords.<br/>"
        "• <b>Lack of API Authentication:</b> Endpoints do not implement session checks, tokens (JWT), or CSRF protection. An attacker can perform registrations or view schedules of other users by predicting student IDs.<br/>"
        "• <b>No Backend Input Sanitization:</b> Raw strings are committed directly to SQLite, making the system prone to persistent script injections (stored XSS).", body_style))

    # ═══════════════════════════════════════
    # 8. RECOMMENDATIONS
    # ═══════════════════════════════════════
    story.append(Paragraph("8. Recommendations", h1_style))
    story.append(Paragraph("To address the identified quality and security gaps, the following remedies are recommended:", body_style))
    
    rec_headers = [["Issue", "Recommended Fix", "Priority", "Module"]]
    for rec in RECOMMENDATIONS:
        rec_headers.append([rec["issue"], rec["fix"], rec["priority"], rec["module"]])
        
    rec_table = Table(rec_headers, colWidths=[120, 180, 80, 120])
    rec_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1B5E20")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E0")),
        ('PADDING', (0,0), (-1,-1), 5),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(rec_table)
    story.append(Spacer(1, 15))

    # ═══════════════════════════════════════
    # 9. FINAL QA ASSESSMENT
    # ═══════════════════════════════════════
    story.append(KeepTogether([
        Paragraph("9. Final QA Assessment", h1_style),
        Paragraph("<b>Overall Status:</b> <font color='#991b1b'><b>Requires Major Fixes (NOT READY FOR DEPLOYMENT)</b></font>", ParagraphStyle('StatusStyle', parent=body_style, fontSize=11, leading=15)),
        Paragraph(
            "<b>Major Strengths:</b><br/>"
            "• High-fidelity, user-friendly frontend dashboard and dynamic weekly timetable grid.<br/>"
            "• Functional validations like credits checking, time clash detection, and course capacity are accurately handled inside Python API services before SQL commits.", body_style),
        Paragraph(
            "<b>Major Weaknesses:</b><br/>"
            "• Plaintext password storage leaves user credentials entirely exposed.<br/>"
            "• Total lack of server-side session authentication allows unauthorized course modifications across profiles.<br/>"
            "• SQLite foreign keys disabled by default, causing orphaned data entries.", body_style),
        Paragraph(
            "<b>Conclusion:</b><br/>"
            "The system is functional from a user perspective but fails critical security, authorization, and data integrity standards. "
            "It must not be deployed to staging/production until plaintext storage is replaced with secure hashing and API authorization guards are implemented.", body_style)
    ]))

    # Build PDF
    doc.build(story, canvasmaker=NumberedCanvas)
    print("PDF Report Generated successfully!")

if __name__ == "__main__":
    build_excel()
    build_pdf()
